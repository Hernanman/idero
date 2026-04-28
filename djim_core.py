"""
DJIM App LOCAL - Extractor de Despacho ARCA/SIM -> DJIM Excel + TXT DNRPA
100% gratuito: no usa OpenAI, Gemini ni Claude.

Uso consola:
  python djim_app_local_gui.py despacho.pdf [DJIM_template.xlsx]

Uso interfaz:
  python djim_app_local_gui.py

Dependencias:
  pip install -r requirements_djim_local.txt
"""

import sys
import os
import json
import re
import shutil
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

# Motor PDF gratuito. Intenta PyMuPDF primero; si no está, cae a pdfplumber.
def extraer_texto_pdf(pdf_path: str) -> str:
    errores = []
    try:
        import fitz  # PyMuPDF
        texto_paginas = []
        with fitz.open(pdf_path) as doc:
            for i, page in enumerate(doc, start=1):
                txt = page.get_text("text") or ""
                texto_paginas.append(f"\n--- PAGINA {i} ---\n{txt}")
        texto = "\n".join(texto_paginas).strip()
        if texto:
            return texto
    except Exception as e:
        errores.append(f"PyMuPDF: {e}")

    try:
        import pdfplumber
        texto_paginas = []
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                txt = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
                texto_paginas.append(f"\n--- PAGINA {i} ---\n{txt}")
        texto = "\n".join(texto_paginas).strip()
        if texto:
            return texto
    except Exception as e:
        errores.append(f"pdfplumber: {e}")

    raise RuntimeError(
        "No pude extraer texto del PDF. Puede ser un PDF escaneado/imagen. "
        "Esta versión gratuita no usa OCR ni IA. Detalle: " + " | ".join(errores)
    )

# Configuración DJIM
DJIM_PRIMERA_FILA_DATOS = 16
ADUANA_MAP = {
    "paso de los libres": "42",
    "rosario": "52",
    "buenos aires": "01",
    "mendoza": "11",
    "cordoba": "36",
    "córdoba": "36",
    "posadas": "26",
}
PAIS_COD_NOMBRE = {
    "203": "BRASIL",
    "225": "URUGUAY",
    "212": "CHILE",
    "221": "PARAGUAY",
    "202": "BOLIVIA",
    "200": "ARGENTINA",
}
MARCA_DNRPA = {
    "IDERO": "V41",
}

# Tabla local DNRPA basada en PLANTILLACODIGOSDRNPA.xlsx.
# Se usa para convertir los sufijos del PDF:
#   AA(IDERO) -> id_marca V41
#   AB(...)   -> id_tipo según comienzo del AB
#   AB(...) + AC(...) -> id_modelo según descripción normalizada
TIPO_DNRPA_POR_AB = {
    "SEMI": "27",
}

MODELO_DNRPA_POR_AB_AC = {
    "SEMI CARGA 3E 3D S3 S3 CS": "002",
    "SEMI CARGA 3E 3D S3 S3 FP": "001",
    "SEMI CARGA CONV 3E 1D2D S12 S12 PC": "004",
    "SEMI CARGA CONV 3E 3D S3 S3 PC": "003",
    "SEMI CARGA CONV 3E 3D S3 S3 VT": "005",
    "SEMI CARGA CONV 3E 3D S3 FR": "008",
    "SEMI CARGA CONV 3E1D 2D S12 S12 VT": "010",
    "SEMI CARGA CONV 3E 1D1D1D S111 VT": "007",
    "SEMI CARGA CONV 3E 1D2D S12 S12 CS": "011",
    "SEMI CARGA CONV 3E ID2D S12": "009",
}

# ---------------- Parser local por reglas ----------------
def limpiar_texto(texto: str) -> str:
    texto = texto.replace("\r", "\n")
    texto = re.sub(r"[ \t]+", " ", texto)
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    return texto

def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def buscar(pattern: str, texto: str, flags=re.I|re.S, group: int = 1, default: str = "") -> str:
    m = re.search(pattern, texto, flags)
    return norm(m.group(group)) if m else default

def buscar_fecha_cerca(etiquetas: List[str], texto: str) -> str:
    for et in etiquetas:
        # etiqueta ... dd/mm/yyyy o dd-mm-yyyy o yyyy-mm-dd
        p = rf"{et}[^\n]{{0,120}}?(\d{{1,2}}[/\-]\d{{1,2}}[/\-]\d{{4}}|\d{{4}}[/\-]\d{{1,2}}[/\-]\d{{1,2}})"
        v = buscar(p, texto)
        if v:
            return formatear_fecha_dnrpa(v)
    # fallback: primera fecha del documento
    v = buscar(r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})", texto)
    return formatear_fecha_dnrpa(v) if v else ""

def extraer_cuit_cerca(etiquetas: List[str], texto: str) -> str:
    for et in etiquetas:
        v = buscar(rf"{et}[^\n]{{0,180}}?(\d{{2}}[- ]?\d{{8}}[- ]?\d)", texto)
        if v:
            return normalizar_cuit(v)
    return ""

def normalizar_cuit(v: str) -> str:
    nums = re.sub(r"\D", "", v or "")
    if len(nums) == 11:
        return f"{nums[:2]}-{nums[2:10]}-{nums[10]}"
    return v

def extraer_rs_cerca(etiquetas: List[str], texto: str) -> str:
    lineas = [norm(x) for x in texto.splitlines() if norm(x)]
    for i, linea in enumerate(lineas):
        for et in etiquetas:
            if re.search(et, linea, re.I):
                # intenta misma línea luego de etiqueta
                limpio = re.sub(et, "", linea, flags=re.I).strip(" :-")
                limpio = re.sub(r"\bCUIT\b.*", "", limpio, flags=re.I).strip(" :-")
                if limpio and len(limpio) > 3 and not re.search(r"^\d", limpio):
                    return limpio[:80]
                # o siguiente línea no numérica
                for j in range(i+1, min(i+4, len(lineas))):
                    cand = re.sub(r"\bCUIT\b.*", "", lineas[j], flags=re.I).strip(" :-")
                    if len(cand) > 3 and not re.match(r"^[\d/\-. ]+$", cand):
                        return cand[:80]
    return ""


def extraer_importador_exportador(texto: str) -> Tuple[str, str]:
    """Detecta el bloque Importador/Exportador en formularios ARCA/SIM.
    En muchos PDFs el texto queda desordenado: la razón social y CUIT aparecen
    antes de la etiqueta 'Importador / Exportador'. Por eso se prioriza la
    línea que contiene '(IVA INS: SI)' y un CUIT.
    """
    lineas = [norm(x) for x in texto.splitlines() if norm(x)]

    # Caso típico del formulario. A veces la razón social y el CUIT quedan
    # en la misma línea; otras veces el CUIT queda en la línea siguiente.
    for i, linea in enumerate(lineas[:80]):
        m = re.search(r"([A-ZÁÉÍÓÚÑ0-9][A-ZÁÉÍÓÚÑ0-9 .,&\-/]{2,80}?)\s*\(\s*IVA\s+INS\s*:\s*SI\s*\)\s*(\d{2}[- ]?\d{8}[- ]?\d)?", linea, re.I)
        if m:
            rs = norm(m.group(1)).upper()
            cuit = normalizar_cuit(m.group(2) or "")
            if not cuit:
                ventana = "\n".join(lineas[i:i+4])
                cuit = normalizar_cuit(buscar(r"(\d{2}[- ]?\d{8}[- ]?\d)", ventana))
            if rs and cuit:
                return rs, cuit

    # Fallback: etiqueta Importador/Exportador y CUIT cercano.
    for i, linea in enumerate(lineas[:100]):
        if re.search(r"IMPORTADOR\s*/?\s*EXPORTADOR|IMPORT\.?/?EXPORTADOR", linea, re.I):
            ventana = "\n".join(lineas[max(0, i-4): min(len(lineas), i+6)])
            cuit = buscar(r"(\d{2}[- ]?\d{8}[- ]?\d)", ventana)
            candidatos = []
            for l in lineas[max(0, i-4): min(len(lineas), i+6)]:
                cand = re.sub(r"\(.*?\)", "", l)
                cand = re.sub(r"\bCUIT\b.*", "", cand, flags=re.I)
                cand = re.sub(r"IMPORTADOR\s*/?\s*EXPORTADOR|IMPORT\.?/?EXPORTADOR|VENDEDOR|DESPACHANTE.*", "", cand, flags=re.I).strip(" :-")
                if len(cand) > 3 and not re.match(r"^[\d*/. -]+$", cand) and "EXPORTADOR" not in cand.upper():
                    candidatos.append(cand)
            if candidatos:
                return norm(candidatos[0]).upper(), normalizar_cuit(cuit)

    return "", ""

def extraer_aduana(texto: str) -> Tuple[str, str]:
    for nombre, codigo in ADUANA_MAP.items():
        if re.search(re.escape(nombre), texto, re.I):
            return nombre.upper(), codigo
    m = re.search(r"ADUANA\s*:?\s*([A-ZÁÉÍÓÚÑ ]{3,50})", texto, re.I)
    if m:
        nombre = norm(m.group(1)).upper()
        return nombre, ADUANA_MAP.get(nombre.lower(), "00")
    return "", "00"

def extraer_despacho(texto: str) -> Tuple[str, str, str]:
    # Ej: 26 042 IC03 001578 U
    m = re.search(r"\b(\d{2})\s+(\d{3})\s+([A-Z]{1,3}\d{2})\s+(\d{4,8})\s+([A-Z])\b", texto)
    if m:
        raw = f"{m.group(1)} {m.group(2)} {m.group(3)} {m.group(4)} {m.group(5)}"
        return raw, m.group(2), m.group(3)
    # Variante pegada
    m = re.search(r"\b(\d{2})(\d{3})([A-Z]{1,3}\d{2})(\d{4,8})([A-Z])\b", texto)
    if m:
        raw = f"{m.group(1)} {m.group(2)} {m.group(3)} {m.group(4)} {m.group(5)}"
        return raw, m.group(2), m.group(3)
    return "", "", ""

def extraer_posicion_sim(texto: str) -> str:
    patterns = [
        r"POSICI[ÓO]N\s+(?:SIM|ARANCELARIA)?[^\n]{0,80}?((?:\d{4}[\. ]?\d{2}[\. ]?\d{2}[\. ]?\d{3}[A-Z]?))",
        r"\b(\d{4}\.\d{2}\.\d{2}\.\d{3}[A-Z])\b",
        r"\b(\d{4}\.\d{2}\.\d{2}\.\d{3})\b",
    ]
    for p in patterns:
        v = buscar(p, texto)
        if v:
            return v.replace(" ", "")
    return ""

def _limpiar_if_lcm(valor: str) -> str:
    """Normaliza la leyenda IF de LCM sin arrastrar texto posterior del PDF."""
    if not valor:
        return ""
    valor = norm(valor).upper()
    # Cortar cualquier texto posterior típico del PDF.
    valor = re.split(r"\bDOCS?\.|\bGUIA/|\bCONOC|\bCAR[ÁA]TULA|\bSUBITEM|\bUNIDAD|\bFOB\b", valor, maxsplit=1, flags=re.I)[0]
    # Quedarse exclusivamente con el IF completo.
    m = re.search(r"(IF\s*-\s*\d{4}\s*-\s*\d{6,10}\s*-\s*[A-Z0-9#\-]+)", valor, re.I)
    if not m:
        return ""
    limpio = re.sub(r"\s+", "", m.group(1)).upper()
    return limpio.strip(" .;:-")


def extraer_lcm(texto: str) -> Tuple[str, str]:
    """Devuelve (lcm_nro, lcm_anio).

    Para TXT DNRPA se sigue usando solo el número del IF.
    Para Excel se guarda además la leyenda completa en cabecera["lcm_leyenda"].
    """
    leyenda = extraer_lcm_leyenda(texto)
    m = re.search(r"IF-(20\d{2})-(\d{6,10})-", leyenda, re.I)
    if m:
        return m.group(2), m.group(1)

    # Fallback: si no se detectó la leyenda completa, buscar cualquier IF del texto.
    m = re.search(r"IF\s*-\s*(20\d{2})\s*-\s*(\d{6,10})\s*-", texto, re.I)
    if m:
        return m.group(2), m.group(1)

    m = re.search(r"LCM[^\n]{0,120}?(\d{6,10})", texto, re.I)
    if m:
        return m.group(1), ""
    return "", ""


def extraer_lcm_leyenda(texto: str) -> str:
    """Extrae la leyenda completa que aparece luego de LICONFIDEMODEL =.

    Soporta los dos formatos que aparecen en los PDFs SIM:
    1) LICONFIDEMODEL = IF-2025-83639995-APN-SSGP#MEC
    2) LICONFIDEMODEL =
       IF-2025-83639995-APN-SSGP#MEC
    """
    # 1) Buscar estrictamente después de la etiqueta LICONFIDEMODEL.
    m_label = re.search(r"LICONFIDEMODEL\s*=", texto, re.I)
    if m_label:
        # Ventana corta posterior a la etiqueta: evita tomar otro IF que no corresponda.
        ventana = texto[m_label.end():m_label.end() + 500]
        valor = _limpiar_if_lcm(ventana)
        if valor:
            return valor

    # 2) Fallback por líneas, útil cuando la extracción de PDF reordena saltos.
    lineas = [norm(x) for x in texto.splitlines() if norm(x)]
    for i, linea in enumerate(lineas):
        if re.search(r"LICONFIDEMODEL", linea, re.I):
            ventana = " ".join(lineas[i:i + 6])
            valor = _limpiar_if_lcm(ventana)
            if valor:
                return valor

    # 3) Fallback general: primer IF completo del documento.
    valor = _limpiar_if_lcm(texto)
    return valor


def extraer_pais(texto: str) -> Tuple[str, str]:
    for cod, nombre in PAIS_COD_NOMBRE.items():
        if re.search(rf"\b{cod}\b[^\n]{{0,40}}{nombre}|{nombre}[^\n]{{0,40}}\b{cod}\b", texto, re.I):
            return cod, nombre
    if re.search(r"BRASIL", texto, re.I):
        return "203", "BRASIL"
    return "", ""

def extraer_valor(etiquetas: List[str], texto: str) -> str:
    for et in etiquetas:
        v = buscar(rf"{et}[^\n]{{0,80}}?([0-9]{{1,3}}(?:[.,][0-9]{{3}})*(?:[.,][0-9]{{2,3}})?|[0-9]+(?:[.,][0-9]{{2,3}})?)", texto)
        if v:
            return normalizar_numero(v)
    return ""

def normalizar_numero(v: str) -> str:
    v = (v or "").strip()
    # 58.419,41 -> 58419.41 ; 58419.41 queda igual
    if "," in v and "." in v:
        if v.rfind(",") > v.rfind("."):
            v = v.replace(".", "").replace(",", ".")
        else:
            v = v.replace(",", "")
    elif "," in v:
        v = v.replace(".", "").replace(",", ".")
    return v


def normalizar_codigo_texto(v: str) -> str:
    """Normaliza textos técnicos de AB/AC para comparar contra la tabla."""
    v = (v or "").upper()
    v = v.replace("_", " ").replace("-", " ")
    v = re.sub(r"[().,/]+", " ", v)
    v = re.sub(r"\s+", " ", v).strip()
    return v


def extraer_sufijo_pdf(texto: str, codigo: str) -> str:
    """Extrae valores AA(...), AB(...), AC(...), ZC(...) tolerando espacios del PDF."""
    letras = list(codigo.upper())
    patron_alt = letras[0] + r"\s*" + letras[1] if len(letras) == 2 else re.escape(codigo)
    patrones = [
        rf"\b{re.escape(codigo)}\s*\(([^)]*)\)",
        rf"\b{patron_alt}\s*\(([^)]*)\)",
    ]
    for patron in patrones:
        m = re.search(patron, texto, re.I | re.S)
        if m:
            return norm(m.group(1)).upper()
    return ""


def extraer_zc_peso_imponible(texto: str) -> str:
    """Extrae el peso imponible desde ZC(047000) y quita ceros a la izquierda."""
    valor = extraer_sufijo_pdf(texto, "ZC")
    if not valor:
        patrones = [
            r"Z\s*C\s*\(\s*([0-9][0-9 .,)\/-]*)\s*\)",
            r"ZC[^0-9]{0,20}([0-9]{4,8})[^\n]{0,100}PESO\s+TOTAL\s+CON\s+CARGA",
            r"([0-9]{4,8})[^\n]{0,100}PESO\s+TOTAL\s+CON\s+CARGA",
        ]
        for patron in patrones:
            m = re.search(patron, texto, re.I | re.S)
            if m:
                valor = m.group(1)
                break
    nums = re.sub(r"\D", "", valor or "")
    if not nums:
        return ""
    return str(int(nums))


def resolver_id_tipo_desde_ab(ab: str) -> str:
    ab_norm = normalizar_codigo_texto(ab)
    for prefijo, codigo in TIPO_DNRPA_POR_AB.items():
        if ab_norm.startswith(normalizar_codigo_texto(prefijo)):
            return codigo
    return ""


def resolver_id_modelo_desde_ab_ac(ab: str, ac: str) -> str:
    combinado = normalizar_codigo_texto(f"{ab} {ac}")
    if not combinado:
        return ""
    if combinado in MODELO_DNRPA_POR_AB_AC:
        return MODELO_DNRPA_POR_AB_AC[combinado]
    tokens_combinado = set(combinado.split())
    mejor_codigo = ""
    mejor_score = -1
    for desc, codigo in MODELO_DNRPA_POR_AB_AC.items():
        tokens_desc = set(desc.split())
        if tokens_desc and tokens_desc.issubset(tokens_combinado):
            score = len(tokens_desc)
            if score > mejor_score:
                mejor_score = score
                mejor_codigo = codigo
    return mejor_codigo

def extraer_vins(texto: str) -> List[str]:
    # VIN/chasis típico: 17 caracteres alfanuméricos, excluyendo I/O/Q. Se priorizan etiquetas AM/VIN/CHASIS.
    candidatos = []
    for p in [
        r"(?:AM\s*\(?|VIN|CHASIS|CHASIS NRO|NRO CHASIS)[^A-Z0-9]{0,20}([A-HJ-NPR-Z0-9]{17})",
        r"\b([A-HJ-NPR-Z0-9]{17})\b",
    ]:
        for m in re.finditer(p, texto, re.I):
            vin = m.group(1).upper()
            if not re.fullmatch(r"\d{17}", vin) and vin not in candidatos:
                candidatos.append(vin)
    return candidatos

def extraer_anios(texto: str) -> Tuple[str, str]:
    anio_modelo = buscar(r"A[ÑN]O\s+MODELO[^\n]{0,50}?(20\d{2})", texto)
    anio_fab = buscar(r"A[ÑN]O\s+(?:FAB|FABRICACI[ÓO]N)[^\n]{0,50}?(20\d{2})", texto)
    # fallback: años 20xx más frecuentes
    years = re.findall(r"\b(20\d{2})\b", texto)
    if not anio_modelo and years:
        anio_modelo = years[-1]
    if not anio_fab:
        anio_fab = anio_modelo
    return anio_modelo, anio_fab

def extraer_marca(texto: str) -> str:
    for marca in MARCA_DNRPA:
        if re.search(rf"\b{re.escape(marca)}\b", texto, re.I):
            return marca
    v = buscar(r"MARCA[^\n]{0,40}?([A-Z0-9][A-Z0-9 .\-/]{2,30})", texto)
    return v.upper() if v else ""


def extraer_total_kg_neto(texto: str) -> str:
    """Extrae el peso imponible desde 'Total Kg. Neto'.

    En los OM/SIM el valor puede aparecer antes, después o en la línea superior
    a la etiqueta. Se evita tomar Peso Guía = 0,000.
    """
    num = r"([0-9]{1,3}(?:\.[0-9]{3})*(?:,[0-9]{1,4})|[0-9]+(?:[.,][0-9]{1,4})?)"

    patrones = [
        rf"{num}\s*TOTAL\s+KG\.?\s+NETO",
        rf"TOTAL\s+KG\.?\s+NETO[^0-9]{{0,120}}{num}",
        rf"{num}\s*TOTAL\s+KGS?\.?\s+NETO",
        rf"TOTAL\s+KGS?\.?\s+NETO[^0-9]{{0,120}}{num}",
    ]
    for patron in patrones:
        m = re.search(patron, texto, re.I | re.S)
        if m:
            # si el patrón tiene un solo grupo, usar grupo 1; si tiene dos por composición, usar el grupo no vacío
            vals = [g for g in m.groups() if g]
            for v in vals:
                val = normalizar_numero(v)
                try:
                    if float(val) > 0:
                        return val
                except Exception:
                    pass

    lineas = [norm(x) for x in texto.splitlines() if norm(x)]
    for i, linea in enumerate(lineas):
        if re.search(r"TOTAL\s+KG\.?\s+NETO|TOTAL\s+KGS?\.?\s+NETO", linea, re.I):
            ventana = " ".join(lineas[max(0, i-3):i+4])
            nums = re.findall(r"[0-9]{1,3}(?:\.[0-9]{3})*(?:,[0-9]{1,4})|[0-9]+(?:[.,][0-9]{1,4})?", ventana)
            for n in nums:
                val = normalizar_numero(n)
                try:
                    if float(val) > 0:
                        return val
                except Exception:
                    pass
    return ""


def extraer_datos_pdf(pdf_path: str) -> dict:
    texto = limpiar_texto(extraer_texto_pdf(pdf_path))
    raw, adu_cod_from_nro, subregimen = extraer_despacho(texto)
    aduana_nombre, aduana_id = extraer_aduana(texto)
    if adu_cod_from_nro and aduana_id == "00":
        aduana_id = str(int(adu_cod_from_nro)) if adu_cod_from_nro.isdigit() else adu_cod_from_nro
    fecha_of = buscar_fecha_cerca(["OFICIALIZ", "FECHA OFIC", "FECHA"], texto)
    fecha_arribo = buscar_fecha_cerca(["ARRIBO", "FECHA ARRIBO"], texto)
    pos_sim = extraer_posicion_sim(texto)
    lcm_leyenda = extraer_lcm_leyenda(texto)
    lcm_nro, lcm_anio = extraer_lcm(texto)
    pais_cod, pais_nombre = extraer_pais(texto)
    marca = extraer_marca(texto)
    aa_pdf = extraer_sufijo_pdf(texto, "AA")
    ab_pdf = extraer_sufijo_pdf(texto, "AB")
    ac_pdf = extraer_sufijo_pdf(texto, "AC")
    zc_peso = extraer_zc_peso_imponible(texto)
    id_marca_resuelto = MARCA_DNRPA.get(aa_pdf or marca, aa_pdf or marca)
    id_tipo_resuelto = resolver_id_tipo_desde_ab(ab_pdf) or buscar(r"\bTIPO[^\n]{0,30}?(\d{1,3})\b", texto)
    id_modelo_resuelto = resolver_id_modelo_desde_ab_ac(ab_pdf, ac_pdf) or "007"
    anio_modelo, anio_fab = extraer_anios(texto)
    vins = extraer_vins(texto)

    importador_rs, importador_cuit = extraer_importador_exportador(texto)
    if not importador_rs:
        importador_rs = extraer_rs_cerca(["IMPORTADOR", "IMPORT./EXPORTADOR", "IMPORTADOR/EXPORTADOR"], texto)
    if not importador_cuit:
        importador_cuit = extraer_cuit_cerca(["IMPORTADOR", "IMPORT./EXPORTADOR", "IMPORTADOR/EXPORTADOR"], texto)

    comprador_rs = extraer_rs_cerca(["COMPRADOR", "CONSIGNATARIO"], texto)
    comprador_cuit = extraer_cuit_cerca(["COMPRADOR", "CONSIGNATARIO"], texto)
    # En despacho DJIM, si no hay comprador declarado distinto, se replica el importador.
    if not comprador_rs or comprador_rs.strip().lower() in {"/ exportador", "exportador"}:
        comprador_rs = importador_rs
    if not comprador_cuit:
        comprador_cuit = importador_cuit
    desp_rs = extraer_rs_cerca(["DESPACHANTE"], texto)
    desp_cuit = extraer_cuit_cerca(["DESPACHANTE"], texto)

    # Si hay varios VIN, genera una línea por VIN. Si no encuentra, genera una para carga manual.
    if not vins:
        vins = [""]

    # Peso imponible: para DJIM/DNRPA se toma de ZC(047000) = PESO TOTAL CON CARGA MAXIMA EN kg.
    # No usar Peso Guía ni Total Kg. Neto para este campo.
    peso_total_kg_neto = extraer_total_kg_neto(texto) or extraer_valor(["TOTAL KG NETO", "PESO BRUTO", "PESO"], texto)
    peso_imponible = zc_peso or peso_total_kg_neto
    peso_por_veh = peso_imponible
    vehiculos = []
    for idx, vin in enumerate(vins, start=1):
        vehiculos.append({
            "orden": str(idx),
            "subitem": str(idx).zfill(4),
            "marca": marca,
            "tipo": id_tipo_resuelto,
            "modelo": id_modelo_resuelto,
            "lcm_tipo": "0",
            "lcm_nro": lcm_nro,
            "lcm_leyenda": lcm_leyenda,
            "lcm_anio": lcm_anio,
            "año_modelo": anio_modelo,
            "año_fab": anio_fab,
            "marca_motor": "",
            "nro_motor": "NO POSEE",
            "marca_chasis": marca,
            "nro_chasis": vin,
            "pais_fab": pais_cod or "203",
            "pais_fab_nombre": pais_nombre or "BRASIL",
            "peso": peso_por_veh,
            "monto_fob": "",
            "id_marca": id_marca_resuelto,
            "id_tipo": id_tipo_resuelto,
            "id_modelo": id_modelo_resuelto,
            "aa_pdf": aa_pdf,
            "ab_pdf": ab_pdf,
            "ac_pdf": ac_pdf,
            "si_bloqueado": "N",
        })

    cabecera = {
        "nro_despacho_raw": raw,
        "aduana_nombre": aduana_nombre,
        "aduana_id": aduana_id,
        "fecha_oficializacion": fecha_of,
        "fecha_arribo": fecha_arribo,
        "subregimen": subregimen,
        "cod_regimen": "20",
        "posicion_sim": pos_sim,
        "importador_rs": importador_rs,
        "importador_cuit": importador_cuit,
        "comprador_rs": comprador_rs,
        "comprador_cuit": comprador_cuit,
        "despachante_rs": desp_rs,
        "despachante_cuit": desp_cuit,
        "pais_procedencia_cod": pais_cod or "203",
        "pais_procedencia_nombre": pais_nombre or "BRASIL",
        "mercaderia_nueva": "si" if re.search(r"NUEV", texto, re.I) else "",
        "mercaderia_usada": "si" if re.search(r"USAD", texto, re.I) else "no",
        "cond_venta": buscar(r"COND(?:ICI[ÓO]N)?\s+VENTA[^\n]{0,40}?([A-Z]{2,3})", texto),
        "fob_total": extraer_valor(["FOB TOTAL", "FOB"], texto),
        "fob_divisa": "DOL" if re.search(r"\bDOL\b|D[ÓO]LAR|USD", texto, re.I) else "",
        "flete_total": extraer_valor(["FLETE"], texto),
        "seguro_total": extraer_valor(["SEGURO"], texto),
        "valor_aduana": extraer_valor(["VALOR ADUANA", "VAL\. ADUANA"], texto),
        "nro_factura": buscar(r"FACTURA[^\n]{0,60}?([A-Z0-9][A-Z0-9\-/]{2,30})", texto),
        "fecha_emision_factura": buscar_fecha_cerca(["EMISI[ÓO]N", "FECHA FACTURA"], texto),
        "total_bultos": extraer_valor(["TOTAL BULTOS", "BULTOS"], texto),
        "peso_bruto": peso_total_kg_neto,
        "peso_imponible": peso_imponible,
        "lcm_tipo": "0",
        "lcm_nro": lcm_nro,
        "lcm_anio": lcm_anio,
        "lcm_leyenda": lcm_leyenda,
        "nro_ref_interna": buscar(r"REF(?:ERENCIA)?\s+INTERNA[^\n]{0,50}?([A-Z0-9\-/]{3,30})", texto),
        "id_marca": id_marca_resuelto,
        "id_tipo": id_tipo_resuelto,
        "id_modelo": id_modelo_resuelto,
        "aa_pdf": aa_pdf,
        "ab_pdf": ab_pdf,
        "ac_pdf": ac_pdf,
        "si_autov": "N",
        "si_certif_electronico": "S" if lcm_nro else "",
    }

    datos = {
        "cabecera": cabecera,
        "vehiculos": vehiculos,
        "liquidacion": {
            "iva_base": extraer_valor(["IVA BASE", "BASE IVA"], texto),
            "iva_porc": buscar(r"IVA[^\n]{0,60}?(10[,.]50|21[,.]00|10\.5|21)", texto),
            "iva_importe": extraer_valor(["IVA IMPORTE", "IVA"], texto),
            "ganancias_porc": buscar(r"GANANCIAS[^\n]{0,60}?(6[,.]00|6)", texto),
            "ganancias_importe": extraer_valor(["GANANCIAS"], texto),
            "arancel_sim": extraer_valor(["ARANCEL SIM", "SIM"], texto),
            "total_pagado": extraer_valor(["TOTAL PAGADO", "TOTAL A PAGAR", "TOTAL"], texto),
        },
        "_auditoria": {
            "motor": "local-regex",
            "advertencia": "Extraccion gratuita por texto/regex. Revisar campos vacios antes de presentar.",
            "campos_vacios": [],
        }
    }
    validar_datos(datos, estricto=False)
    datos["_auditoria"]["campos_vacios"] = campos_vacios_importantes(datos)
    return datos

def campos_vacios_importantes(datos: dict) -> List[str]:
    checks = [
        ("cabecera.nro_despacho_raw", datos.get("cabecera", {}).get("nro_despacho_raw")),
        ("cabecera.fecha_oficializacion", datos.get("cabecera", {}).get("fecha_oficializacion")),
        ("cabecera.importador_cuit", datos.get("cabecera", {}).get("importador_cuit")),
        ("cabecera.posicion_sim", datos.get("cabecera", {}).get("posicion_sim")),
        ("cabecera.lcm_nro", datos.get("cabecera", {}).get("lcm_nro")),
    ]
    for i, v in enumerate(datos.get("vehiculos", []), start=1):
        checks.append((f"vehiculos[{i}].nro_chasis", v.get("nro_chasis")))
    return [k for k, v in checks if not v]

def validar_datos(datos: dict, estricto: bool = True) -> None:
    if not isinstance(datos, dict):
        raise ValueError("La extracción no devolvió un diccionario.")
    if "cabecera" not in datos or "vehiculos" not in datos:
        raise ValueError("Faltan claves obligatorias: cabecera / vehiculos.")
    if not isinstance(datos["cabecera"], dict):
        raise ValueError("cabecera debe ser objeto JSON.")
    if not isinstance(datos["vehiculos"], list):
        raise ValueError("vehiculos debe ser una lista.")
    obligatorios = ["nro_despacho_raw", "fecha_oficializacion"] if estricto else []
    faltan = [c for c in obligatorios if not datos["cabecera"].get(c)]
    if faltan:
        raise ValueError("Faltan campos obligatorios: " + ", ".join(faltan))

# ─── Formateo del número de despacho para el TXT ─────────────────────────────

def formatear_nro_despacho_txt(raw: str) -> str:
    """
    Convierte '26 042 IC03 001578 U' → 'IC03001578U/26'
    """
    partes = raw.strip().split()
    if len(partes) >= 5:
        anio, adu, tipo, nro, dc = partes[0], partes[1], partes[2], partes[3], partes[4]
        return f"{tipo}{nro}{dc}/{anio}"
    return raw.replace(" ", "")


def formatear_fecha_excel(fecha_str: str):
    """Convierte dd/mm/yyyy a datetime para Excel."""
    try:
        return datetime.strptime(fecha_str, "%d/%m/%Y")
    except Exception:
        return fecha_str


# ─── Llenar el DJIM Excel ─────────────────────────────────────────────────────

def _celda_real_escritura(sh, coord: str) -> str:
    """Devuelve la celda superior izquierda si coord pertenece a un rango combinado."""
    for rango in sh.merged_cells.ranges:
        if coord in rango:
            return rango.start_cell.coordinate
    return coord

def _set_excel(sh, coord: str, valor):
    sh[_celda_real_escritura(sh, coord)] = valor

def _buscar_celda_texto(sh, patron: str, max_row: int = 60):
    rx = re.compile(patron, re.I)
    for row in sh.iter_rows(min_row=1, max_row=max_row):
        for c in row:
            if isinstance(c.value, str) and rx.search(c.value):
                return c
    return None

def _clear_value_in_region(sh, texto_objetivo: str, min_row: int = 1, max_row: int = 14):
    """Borra valores exactos repetidos en la zona superior del formulario."""
    if not texto_objetivo:
        return
    objetivo = norm(str(texto_objetivo)).upper()
    for row in range(min_row, min(max_row, sh.max_row) + 1):
        for col in range(1, sh.max_column + 1):
            val = sh.cell(row, col).value
            if isinstance(val, str) and norm(val).upper() == objetivo:
                _set_excel(sh, sh.cell(row, col).coordinate, "")


def _write_once(sh, row: int, col: int, valor: str):
    if valor:
        _set_excel(sh, sh.cell(row, col).coordinate, valor)


def _clear_value_in_block(sh, fila: int, texto_objetivo: str):
    if not texto_objetivo:
        return
    objetivo = norm(str(texto_objetivo)).upper()
    for row in (fila, fila + 1):
        if row < 1 or row > sh.max_row:
            continue
        for col in range(1, sh.max_column + 1):
            val = sh.cell(row, col).value
            if isinstance(val, str) and norm(val).upper() == objetivo:
                _set_excel(sh, sh.cell(row, col).coordinate, "")


def _find_row_containing(sh, patron: str, max_row: int = 20) -> int | None:
    rx = re.compile(patron, re.I)
    for row in range(1, min(max_row, sh.max_row) + 1):
        textos = " ".join(str(sh.cell(row, col).value or "") for col in range(1, sh.max_column + 1))
        if rx.search(textos):
            return row
    return None


def _limpiar_exportador_y_repetidos(sh, fila: int, razon_social: str, cuit: str):
    objetivos = {"/ EXPORTADOR", "EXPORTADOR", norm(razon_social).upper(), norm(cuit).upper()}
    for row in range(max(1, fila - 1), min(sh.max_row, fila + 3) + 1):
        for col in range(1, sh.max_column + 1):
            val = sh.cell(row, col).value
            if isinstance(val, str):
                val_norm = norm(val).upper()
                if val_norm in objetivos or re.fullmatch(r"/?\s*EXPORTADOR", val_norm):
                    _set_excel(sh, sh.cell(row, col).coordinate, "")


def _completar_identificacion_por_etiqueta(sh, patron_etiqueta: str, razon_social: str, cuit: str):
    """Completa Importador/Comprador sin pisar las leyendas inferiores."""
    if not razon_social and not cuit:
        return False

    es_comprador = re.search(r"COMPRADOR", patron_etiqueta, re.I) is not None
    fila_preferida = 9 if es_comprador else 7
    col_nombre_preferida = 6 if es_comprador else 4
    col_cuit_preferida = 12

    fila_detectada = _find_row_containing(sh, patron_etiqueta, max_row=20)
    fila = fila_preferida
    if fila_detectada and abs(fila_detectada - fila_preferida) > 2:
        fila = fila_detectada

    _limpiar_exportador_y_repetidos(sh, fila, razon_social, cuit)
    _set_excel(sh, sh.cell(fila, col_nombre_preferida).coordinate, razon_social)
    _set_excel(sh, sh.cell(fila, col_cuit_preferida).coordinate, cuit)

    for col in range(1, sh.max_column + 1):
        val = sh.cell(fila, col).value
        if isinstance(val, str) and re.fullmatch(r"/?\s*Exportador", norm(val), re.I):
            _set_excel(sh, sh.cell(fila, col).coordinate, "")
    return True


def _set_si_posible(sh, row: int, col: int, valor):
    if row < 1 or col < 1 or row > sh.max_row or col > sh.max_column:
        return
    try:
        _set_excel(sh, sh.cell(row, col).coordinate, valor)
    except Exception:
        pass


def _forzar_identidad_template_djim(sh, razon_social: str, cuit: str, tipo: str):
    """Compatibilidad: delega al completador único para evitar duplicados."""
    patron = r"IMPORTADOR" if tipo == "importador" else r"COMPRADOR\s+DECLARADO"
    return _completar_identificacion_por_etiqueta(sh, patron, razon_social, cuit)


def llenar_djim_excel(datos: dict, template_path: str, output_path: str):
    """Copia el template DJIM y completa con los datos extraídos."""
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    sh = wb["ANVERSO"]

    cab = datos["cabecera"]
    vehs = datos["vehiculos"]

    # ── Cabecera ────────────────────────────────────────────────────
    # Nro despacho (mostrar tal cual aparece en el PDF)
    sh["E3"] = cab.get("nro_despacho_raw", "")
    
    # Fecha de nacionalización
    fecha_of = cab.get("fecha_oficializacion", "")
    sh["J3"] = formatear_fecha_excel(fecha_of)
    sh["J3"].number_format = "DD/MM/YYYY"

    # Código régimen e importación
    sh["L3"] = cab.get("cod_regimen", "20")
    sh["O3"] = cab.get("posicion_sim", "")

    # Importador / comprador: se completa por búsqueda de etiqueta para no duplicar
    # la razón social en varias celdas del template.
    _completar_identificacion_por_etiqueta(
        sh, r"IMPORTADOR", cab.get("importador_rs", ""), cab.get("importador_cuit", "")
    )
    _completar_identificacion_por_etiqueta(
        sh, r"COMPRADOR\s+DECLARADO", cab.get("comprador_rs", ""), cab.get("comprador_cuit", "")
    )

    # País procedencia
    sh["E11"] = cab.get("pais_procedencia_cod", "203")
    sh["J11"] = cab.get("mercaderia_nueva", "si")
    sh["L11"] = cab.get("mercaderia_usada", "no")

    # Aduana
    sh["C31"] = cab.get("aduana_nombre", "")
    sh["C35"] = cab.get("aduana_nombre", "")

    # ── Vehículos ────────────────────────────────────────────────────
    for i, veh in enumerate(vehs):
        fila = DJIM_PRIMERA_FILA_DATOS + i
        sh[f"B{fila}"] = veh.get("id_marca", cab.get("id_marca", ""))
        sh[f"C{fila}"] = veh.get("id_tipo", cab.get("id_tipo", veh.get("tipo", "")))
        sh[f"D{fila}"] = veh.get("id_modelo", cab.get("id_modelo", veh.get("modelo", "")))
        sh[f"E{fila}"] = veh.get("lcm_leyenda") or cab.get("lcm_leyenda") or veh.get("lcm_nro", "")
        sh[f"F{fila}"] = veh.get("año_modelo", "")
        sh[f"G{fila}"] = veh.get("año_fab", "")
        sh[f"H{fila}"] = veh.get("marca_motor", "")
        sh[f"I{fila}"] = veh.get("nro_motor", "")
        sh[f"J{fila}"] = veh.get("marca_chasis", "")
        sh[f"K{fila}"] = veh.get("nro_chasis", "")
        sh[f"L{fila}"] = veh.get("pais_fab_nombre", veh.get("pais_fab", ""))
        sh[f"M{fila}"] = veh.get("peso", "")

    # Fecha actual para el formulario
    sh["B37"] = datetime.today().strftime("%d/%m/%Y") if "B37" in [c.coordinate for r in sh for c in r] else None

    wb.save(output_path)
    print(f"  ✓ DJIM Excel guardado: {output_path}")


# ─── Generar TXT DNRPA ────────────────────────────────────────────────────────

def formatear_fecha_dnrpa(fecha_str: str) -> str:
    """
    Convierte fecha a DD/MM/YYYY con ceros obligatorios.
    Acepta: dd/mm/yyyy, d/m/yyyy, yyyy-mm-dd, etc.
    """
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%-d/%-m/%Y", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(fecha_str.strip(), fmt)
            return dt.strftime("%d/%m/%Y")   # siempre 2 dígitos día y mes
        except ValueError:
            continue
    # Último intento: parsear partes manualmente
    try:
        partes = fecha_str.replace("-", "/").split("/")
        if len(partes) == 3:
            p0, p1, p2 = partes
            # Detectar si está en formato yyyy/mm/dd
            if len(p0) == 4:
                return f"{p1.zfill(2)}/{p2.zfill(2)}/{p0}"
            else:
                return f"{p0.zfill(2)}/{p1.zfill(2)}/{p2}"
    except Exception:
        pass
    return fecha_str   # devolver tal cual si no se puede parsear


def _csv_line(campos: list) -> str:
    """Genera una línea con campos entre comillas dobles separados por punto y coma."""
    return ";".join(f'"{v}"' for v in campos)


def generar_txt_dnrpa(datos: dict, output_path: str):
    """
    Genera el TXT para importar en DNRPA.
    Formato: campos entre comillas dobles, separados por punto y coma (;).
    Fechas siempre con ceros: DD/MM/YYYY.
    Sin línea de encabezados (el sistema DNRPA no los usa).
    """
    cab = datos["cabecera"]
    vehs = datos["vehiculos"]

    nro_raw = cab.get("nro_despacho_raw", "")
    nro_txt = formatear_nro_despacho_txt(nro_raw)

    aduana_nombre = cab.get("aduana_nombre", "").lower()
    id_adu = ADUANA_MAP.get(aduana_nombre, cab.get("aduana_id", "00"))

    # Fecha con ceros obligatorios
    fecha_of_str = cab.get("fecha_oficializacion", "")
    fecha_of_fmt = formatear_fecha_dnrpa(fecha_of_str)

    cant_lineas = str(len(vehs))

    # ── Línea 1: Cabecera del despacho (20 campos) ──────────────────
    linea_cab = [
        id_adu,                                  # 01 ID_ADU
        nro_txt,                                 # 02 NRO_DESPACHO
        "0",                                     # 03 ORDEN_DESPACHO
        "12",                                    # 04 TIPO DOC IMP (CUIT=12)
        cab.get("importador_cuit", ""),          # 05 NRO DOC IMP
        "12",                                    # 06 TIPO DOC COMP
        cab.get("comprador_cuit", ""),           # 07 NRO DOC COM
        "12",                                    # 08 TIPO DOC DESP
        cab.get("despachante_cuit", ""),         # 09 NRO DOC DESP
        "S",                                     # 10 ID_REGIMEN
        fecha_of_fmt,                            # 11 FECHA OFIC (DD/MM/YYYY con ceros)
        cab.get("pais_procedencia_cod", "203"),  # 12 PAÍS PROC
        cant_lineas,                             # 13 CANT.LÍNEAS
        cab.get("si_autov", "N"),                # 14 SI_AUTOV
        cab.get("si_certif_electronico", "S"),   # 15 SI_CERTIF_ELECTRONICO
        "",                                      # 16 TIPO DOC EMB
        "",                                      # 17 NRO DOC EMB
        "",                                      # 18 FECHA PLAZA
        "",                                      # 19 NRO FRANQ
        "",                                      # 20 (campo extra observado en el TXT)
    ]

    lineas = [_csv_line(linea_cab)]

    # ── Línea por vehículo (19 campos) ─────────────────────────────
    for veh in vehs:
        linea_veh = [
            id_adu,                                          # 01 ID_ADU
            nro_txt,                                         # 02 NRO_DESPACHO
            "0",                                             # 03 ORDEN_DESPACHO
            veh.get("orden", "1"),                           # 04 ORDEN
            veh.get("id_marca", cab.get("id_marca", "")),   # 05 ID_MARCA
            veh.get("tipo", cab.get("id_tipo", "")),         # 06 ID_TIPO
            veh.get("id_modelo", cab.get("id_modelo", "007")),   # 07 ID_MODELO
            veh.get("lcm_tipo", cab.get("lcm_tipo", "0")),  # 08 LCM_TIPO
            veh.get("lcm_nro", cab.get("lcm_nro", "")),     # 09 LCM_NRO
            veh.get("lcm_anio", cab.get("lcm_anio", "")),   # 10 LCM_AÑO (año del IF)
            veh.get("año_modelo", ""),                       # 11 AÑO_MODELO
            veh.get("año_fab", ""),                          # 12 AÑO_FAB
            "000",                                           # 13 ID_MARCA_MOTOR
            veh.get("nro_motor", "NO POSEE"),                # 14 NRO MOTOR
            veh.get("id_marca", cab.get("id_marca", "")),   # 15 ID_MARCA_CHASIS
            veh.get("nro_chasis", ""),                       # 16 NRO CHASIS
            veh.get("pais_fab", "203"),                      # 17 PAIS FAB
            veh.get("peso", ""),                             # 18 PESO IMP
            veh.get("si_bloqueado", "N"),                    # 19 SI_BLOQUEADO
        ]
        lineas.append(_csv_line(linea_veh))

    # Sin encabezados — el sistema DNRPA no los espera
    with open(output_path, "w", encoding="utf-8", newline="\n") as f:
        for linea in lineas:
            f.write(linea + "\n")

    print(f"  ✓ TXT DNRPA guardado: {output_path}")


# ─── Resumen en consola ───────────────────────────────────────────────────────

def imprimir_resumen(datos: dict):
    cab = datos["cabecera"]
    vehs = datos["vehiculos"]
    print("\n" + "="*60)
    print("  RESUMEN DEL DESPACHO EXTRAÍDO")
    print("="*60)
    print(f"  Despacho:      {cab.get('nro_despacho_raw','')}")
    print(f"  Aduana:        {cab.get('aduana_nombre','')} (ID: {cab.get('aduana_id','')})")
    print(f"  Oficializ.:    {cab.get('fecha_oficializacion','')}")
    print(f"  Importador:    {cab.get('importador_rs','')} - {cab.get('importador_cuit','')}")
    print(f"  Posición SIM:  {cab.get('posicion_sim','')}")
    print(f"  FOB Total:     USD {cab.get('fob_total','')}")
    print(f"  Vehículos:     {len(vehs)}")
    for v in vehs:
        print(f"    [{v.get('orden','')}] {v.get('marca','')} | Chasis: {v.get('nro_chasis','')} | Año fab: {v.get('año_fab','')} | Peso: {v.get('peso','')} kg")
    print("="*60 + "\n")





def procesar_djim_web(pdf_path: str, output_dir: str, template_path: str | None = None) -> dict:
    """Procesa un PDF y genera JSON/TXT/XLSX en output_dir. Retorna rutas y datos."""
    output_dir_p = Path(output_dir)
    output_dir_p.mkdir(parents=True, exist_ok=True)

    base = Path(pdf_path).stem
    output_xlsx = output_dir_p / f"DJIM_{base}.xlsx"
    output_txt = output_dir_p / f"DJIM_{base}.txt"
    output_json = output_dir_p / f"DJIM_{base}_datos.json"

    datos = extraer_datos_pdf(pdf_path)
    validar_datos(datos, estricto=False)

    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)

    xlsx_generado = None
    if template_path and os.path.exists(template_path):
        llenar_djim_excel(datos, template_path, str(output_xlsx))
        xlsx_generado = str(output_xlsx)

    generar_txt_dnrpa(datos, str(output_txt))

    return {
        "datos": datos,
        "json_path": str(output_json),
        "txt_path": str(output_txt),
        "xlsx_path": xlsx_generado,
        "campos_vacios": datos.get("_auditoria", {}).get("campos_vacios", []),
    }
